"""Excel workbook formatting helpers."""

import logging
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ── Column type mappings ───────────────────────────────────────────────────────
_DOLLAR_WHOLE = {
    'CUR_MKT_CAP', 'CASH_AND_EQUIVS', 'BS_ST_BORROW', 'BS_ST_LEASE_LIAB',
    'ST_DEBT', 'BS_LT_BORROW', 'LT_LEASES', 'LT_DEBT', 'CFO_Q', 'CFO_TTM',
    'FCF_Q', 'FCF_TTM', 'AVAT_100d', 'AVAT_20d', 'EQY_FLOAT1', 'EQY_FLOAT2',
    'FloatValue_preFile', 'FloatValue_postFile', 'FloatValue_last60d',
}

_DOLLAR_CENTS = {
    'PX_LAST', 'INTERVAL_HIGH1', 'INTERVAL_HIGH2', 'INTERVAL_HIGH_60D',
}

_PERCENT_DECIMAL = {          # stored as 0.15 → display 15%
    'AVAT_100d_Burn_Q', 'AVAT_20d_Burn_Q',
    'AVAT_100d_Burn_TTM', 'AVAT_20d_Burn_TTM',
}

_PERCENT_DIVIDE_100 = {       # stored as 15.0 → divide by 100 → display 15%
    '3MO_CALL_IMP_VOL', '12MO_CALL_IMP_VOL', 'VOLATILITY_90D',
}

_DATE_COLS = {
    'MOST_RECENT_PERIOD_END_DT', 'LATEST_ANN_DT_QTRLY', 'OFFERING_PRELIM_FILING_DT',
    'LATEST_ANN_DT_ANNUAL', '10-K_Date', '10-K_Date_minus60d',
}

_INTEGER_COLS = {'Burn_Q', 'Burn_TTM'}


def format_excel_columns(filepath: str) -> str:
    """
    Apply number/date/currency formatting to an Excel file in-place.

    Parameters
    ----------
    filepath : str
        Path to the .xlsx file.

    Returns
    -------
    str
        The same filepath (for chaining).
    """
    wb = load_workbook(filepath)
    ws = wb.active

    # Map header names → column indices
    headers = {
        cell.value: col_idx
        for col_idx, cell in enumerate(ws[1], start=1)
        if cell.value
    }

    for col_name, col_idx in headers.items():
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col_idx)
            if cell.value in (None, ''):
                continue

            if col_name in _DOLLAR_CENTS:
                cell.number_format = '$#,##0.00_);[Red]($#,##0.00)'

            elif col_name in _DOLLAR_WHOLE:
                cell.number_format = '$#,##0_);[Red]($#,##0)'

            elif col_name == 'shelf_limit':
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0_);[Red]($#,##0)'

            elif col_name in _PERCENT_DECIMAL:
                cell.number_format = '0.00%'

            elif col_name in _PERCENT_DIVIDE_100:
                try:
                    val = float(str(cell.value).replace('%', '').strip())
                    cell.value = val / 100
                    cell.number_format = '0.00%'
                except (ValueError, TypeError):
                    pass

            elif col_name in _DATE_COLS:
                cell.number_format = 'MM/DD/YYYY'

            elif col_name in _INTEGER_COLS:
                try:
                    if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                        cell.value = int(cell.value)
                    cell.number_format = '#,##0'
                except (ValueError, TypeError):
                    pass

    wb.save(filepath)
    logger.info(f"Formatted: {filepath}")
    return filepath
