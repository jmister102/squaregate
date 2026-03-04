import blpapi
import pandas as pd
import numpy as np
from datetime import datetime
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def apply_filters(df):
    """
    Apply all filter tests and return DataFrame with test columns.
    Returns dict of {test_name: test_result_series}
    """
    filters = {}

    # Filter 1: Market cap < $75M (PASS if less than 75M)
    filters['test_float_lt_75M'] = df['FloatValue_last60d'] < 75_000_000

    # Filter 2: Volume check (PASS if either volume > 50k)
    filters['test_adv_gt_50k'] = (df['AVAT_100d'] > 50_000) | (df['AVAT_20d'] > 50_000)

    # Filter: < 24 months runway (cash/cash equiv)/free cash flow
    # Does this need to be burn_q *OR* burn_ttm < 24 or just burn_q??
    df['Burn_Q'] = pd.to_numeric(df['Burn_Q'], errors='coerce')
    df['Burn_TTM'] = pd.to_numeric(df['Burn_TTM'], errors='coerce')
    filters['test_lt_24m_runway'] = (
            (df['Burn_Q'].between(0, 24, inclusive='neither')) |
            (df['Burn_TTM'].between(0, 24, inclusive='neither'))
    )
    # Filter: ADV/daily burn < 35% (PASS if cash flow positive OR ratio < 35%)
    daily_burn = df['CFO_Q'] / 63
    filters['test_adv_div_dailyburn_lt_35pct'] = (
        (df['CFO_Q'] >= 0) |  # PASS if cash flow positive (not burning)
        (df['AVAT_100d'] / daily_burn < 0.35)  # PASS if burning but ratio < 35%
    )
    return filters


def format_excel_columns(filepath):
    """
    Format Excel columns with proper types (currency, percentage, date, integer).

    Parameters:
    -----------
    filepath : str
        Path to the Excel file to format
    """
    # Define column types - using lowercase to match DataFrame columns
    dollar_columns = [
        'CUR_MKT_CAP', 'CASH_AND_EQUIVS', 'BS_ST_BORROW',
        'BS_ST_LEASE_LIAB', 'ST_DEBT', 'BS_LT_BORROW', 'LT_LEASES',
        'LT_DEBT', 'CFO_Q', 'CFO_TTM', 'FCF_Q', 'FCF_TTM',
        'AVAT_100d', 'AVAT_20d',
        'EQY_FLOAT1', 'EQY_FLOAT2', 'FloatValue_preFile',
        'FloatValue_postFile', 'FloatValue_last60d'
    ]

    dollarcent_columns = [
        'PX_LAST',
        'INTERVAL_HIGH1', 'INTERVAL_HIGH2', 'INTERVAL_HIGH_60D'
    ]

    percentage_columns = [
        'AVAT_100d_Burn_Q', 'AVAT_20d_Burn_Q', 'AVAT_100d_Burn_TTM', 'AVAT_20d_Burn_TTM'
    ]


    percentage_columns_div100 = [
        '3MO_CALL_IMP_VOL', '12MO_CALL_IMP_VOL', 'VOLATILITY_90D'
    ]

    date_columns = [
        'MOST_RECENT_PERIOD_END_DT', 'LATEST_ANN_DT_QTRLY', 'OFFERING_PRELIM_FILING_DT',
        'LATEST_ANN_DT_ANNUAL', '10-K_Date', '10-K_Date_minus60d'
    ]

    integer_columns = [
        'Burn_Q', 'Burn_TTM'
    ]

    # Load the workbook
    wb = load_workbook(filepath)
    ws = wb.active

    # Get header row to map column names to indices
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value] = col_idx

        # Apply formatting to each column type
    for col_name, col_idx in headers.items():
        col_letter = ws.cell(1, col_idx).column_letter

        # Dollar formatting:
        if col_name in dollarcent_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is not None and cell.value != '':
                    cell.number_format = '$#,##0.00_);[Red]($#,##0.00)'

        # Dollar.cent formatting:
        elif col_name in dollar_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is not None and cell.value != '':
                    cell.number_format = '$#,##0_);[Red]($#,##0)'

        elif col_name == 'shelf_limit':
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value not in (None, ''):
                    # Only format as currency if it's a number
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0_);[Red]($#,##0)'
                    # Otherwise leave as text ('Unlimited' or '?')

        # Percentage formatting: 0.00%
        elif col_name in percentage_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is not None and cell.value != '':
                    # Convert decimal to percentage if needed (0.15 -> 15%)
                    cell.number_format = '0.00%'

        # Percentage formatting: 0.00%
        elif col_name in percentage_columns_div100:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value not in (None, ''):
                    try:
                        # Handle string values with % sign already
                        val = str(cell.value).replace('%', '').strip()
                        cell.value = float(val) / 100
                        cell.number_format = '0.00%'
                    except (ValueError, TypeError):
                        # Leave as-is if conversion fails
                        continue

        # Date formatting: MM/DD/YYYY
        elif col_name in date_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is not None and cell.value != '':
                    cell.number_format = 'MM/DD/YYYY'

        elif col_name in integer_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value not in (None, ''):
                    try:
                        # Convert float to int for display (handles 24.0 -> 24)
                        if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                            cell.value = int(cell.value)
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        pass

    # Save the formatted workbook
    wb.save(filepath)
    logger.info(f"Applied formatting to {filepath}")
    return filepath


def create_fail_reason(df, test_cols):
    """Create a string listing all failed tests for each row."""
    # Efficient vectorized approach
    failed = ~df[test_cols]  # Invert to get failures

    fail_reasons = []
    for idx in df.index:
        failed_tests = [col for col in test_cols if failed.loc[idx, col]]
        fail_reasons.append(', '.join(failed_tests))

    return fail_reasons


class BloombergDataCollector:
    """Class to collect Bloomberg data for multiple tickers"""

    def __init__(self, host='localhost', port=8194):
        """Initialize Bloomberg API connection"""
        self.host = host
        self.port = port
        self.session = None

    def connect(self):
        """Establish connection to Bloomberg API"""
        try:
            sessionOptions = blpapi.SessionOptions()
            sessionOptions.setServerHost(self.host)
            sessionOptions.setServerPort(self.port)

            self.session = blpapi.Session(sessionOptions)

            if not self.session.start():
                logger.error("Failed to start Bloomberg session")
                return False

            if not self.session.openService("//blp/refdata"):
                logger.error("Failed to open //blp/refdata service")
                return False

            logger.info("Successfully connected to Bloomberg API")
            return True

        except Exception as e:
            logger.error(f"Connection error: {str(e)}")
            return False

    def disconnect(self):
        """Close Bloomberg API connection"""
        if self.session:
            self.session.stop()
            logger.info("Disconnected from Bloomberg API")

    def get_reference_data(self, tickers, fields, overrides=None):
        """
        Fetch reference data for given tickers and fields

        Parameters:
        -----------
        tickers : list
            List of ticker symbols (e.g., ['AAPL US Equity', 'MSFT US Equity'])
        fields : list
            List of Bloomberg field names
        overrides : dict or list of dicts (optional)
            Dictionary of field overrides or list of override dicts
            Example: {'CALC_INTERVAL': '100D', 'END_DATE_OVERRIDE': '20231215'}

        Returns:
        --------
        pandas.DataFrame with tickers as rows and fields as columns
        """
        if not self.session:
            logger.error("Session not established. Call connect() first.")
            return None

        try:
            refDataService = self.session.getService("//blp/refdata")
            request = refDataService.createRequest("ReferenceDataRequest")

            # Add tickers
            for ticker in tickers:
                request.append("securities", ticker)

            # Add fields
            for field in fields:
                request.append("fields", field)

            # Add overrides if provided
            if overrides:
                overridesElement = request.getElement("overrides")
                for key, value in overrides.items():
                    override = overridesElement.appendElement()
                    override.setElement("fieldId", key)
                    override.setElement("value", str(value))
                logger.info(f"Applied {len(overrides)} overrides")

            logger.info(f"Sending request for {len(tickers)} tickers and {len(fields)} fields")
            self.session.sendRequest(request)

            # Process response
            data_dict = {ticker: {} for ticker in tickers}

            while True:
                event = self.session.nextEvent(500)

                if event.eventType() == blpapi.Event.RESPONSE or \
                        event.eventType() == blpapi.Event.PARTIAL_RESPONSE:

                    for msg in event:
                        securityDataArray = msg.getElement("securityData")

                        for i in range(securityDataArray.numValues()):
                            securityData = securityDataArray.getValueAsElement(i)
                            ticker = securityData.getElementAsString("security")

                            fieldData = securityData.getElement("fieldData")

                            for field in fields:
                                try:
                                    if fieldData.hasElement(field):
                                        value = fieldData.getElement(field).getValue()
                                        data_dict[ticker][field] = value
                                    else:
                                        data_dict[ticker][field] = None
                                except Exception as e:
                                    logger.warning(f"Error getting {field} for {ticker}: {str(e)}")
                                    data_dict[ticker][field] = None

                            # Check for field exceptions
                            if securityData.hasElement("fieldExceptions"):
                                fieldExceptions = securityData.getElement("fieldExceptions")
                                for j in range(fieldExceptions.numValues()):
                                    fieldException = fieldExceptions.getValueAsElement(j)
                                    fieldId = fieldException.getElementAsString("fieldId")
                                    logger.warning(f"Field exception for {ticker} - {fieldId}")

                if event.eventType() == blpapi.Event.RESPONSE:
                    break

            # Convert to DataFrame
            df = pd.DataFrame.from_dict(data_dict, orient='index')
            df.index.name = 'Ticker'

            logger.info(f"Successfully retrieved data for {len(df)} tickers")
            return df

        except Exception as e:
            logger.error(f"Error fetching data: {str(e)}")
            return None

    def get_historical_data(self, tickers, fields, start_date, end_date=None):
        """
        Fetch historical data for given tickers and fields

        Parameters:
        -----------
        tickers : list
            List of ticker symbols
        fields : list
            List of Bloomberg field names
        start_date : str
            Start date in YYYYMMDD format
        end_date : str (optional)
            End date in YYYYMMDD format. If None, uses start_date (single date)

        Returns:
        --------
        pandas.DataFrame with tickers as rows and fields as columns
        """
        if not self.session:
            logger.error("Session not established. Call connect() first.")
            return None

        try:
            refDataService = self.session.getService("//blp/refdata")
            request = refDataService.createRequest("HistoricalDataRequest")

            # Add tickers
            for ticker in tickers:
                request.append("securities", ticker)

            # Add fields
            for field in fields:
                request.append("fields", field)

            # Set dates
            request.set("startDate", start_date)
            request.set("endDate", end_date if end_date else start_date)

            logger.info(f"Sending historical request for {len(tickers)} tickers, date: {start_date}")
            self.session.sendRequest(request)

            # Process response
            data_dict = {ticker: {} for ticker in tickers}

            while True:
                event = self.session.nextEvent(500)

                if event.eventType() == blpapi.Event.RESPONSE or \
                        event.eventType() == blpapi.Event.PARTIAL_RESPONSE:

                    for msg in event:
                        securityData = msg.getElement("securityData")
                        ticker = securityData.getElementAsString("security")

                        fieldDataArray = securityData.getElement("fieldData")

                        # Get the last (most recent) data point
                        if fieldDataArray.numValues() > 0:
                            fieldData = fieldDataArray.getValueAsElement(fieldDataArray.numValues() - 1)

                            for field in fields:
                                try:
                                    if fieldData.hasElement(field):
                                        value = fieldData.getElement(field).getValue()
                                        data_dict[ticker][field] = value
                                    else:
                                        data_dict[ticker][field] = None
                                except Exception as e:
                                    logger.warning(f"Error getting {field} for {ticker}: {str(e)}")
                                    data_dict[ticker][field] = None
                        else:
                            # No data for this date
                            for field in fields:
                                data_dict[ticker][field] = None

                if event.eventType() == blpapi.Event.RESPONSE:
                    break

            # Convert to DataFrame
            df = pd.DataFrame.from_dict(data_dict, orient='index')
            df.index.name = 'Ticker'

            logger.info(f"Successfully retrieved historical data for {len(df)} tickers")
            return df

        except Exception as e:
            logger.error(f"Error fetching historical data: {str(e)}")
            return None


def main():
    """Example usage"""

    tickers = pd.read_csv('C:/dev/more_tickers.csv', header=None)[0].dropna().tolist()

    # Calculate dates for overrides
    from datetime import timedelta
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y%m%d')
    sixty_days_ago = (datetime.now() - timedelta(days=60)).strftime('%Y%m%d')

    # Define standard fields (no overrides needed)
    standard_fields = [
        'NAME',
        'INDUSTRY_GROUP',
        'PX_LAST',
        # 'EQY_INIT_PO_DT',
        '3MO_CALL_IMP_VOL',  # 3MO_CALL_IMP_VOL
        '12MO_CALL_IMP_VOL',  # 12MO_CALL_IMP_VOL
        'VOLATILITY_90D',
        'CUR_MKT_CAP',
        'ST_DEBT',
        'LT_DEBT',
        'MOST_RECENT_PERIOD_END_DT',
        'LATEST_ANN_DT_QTRLY',
        'OFFERING_PRELIM_FILING_DT',
        'LATEST_ANN_DT_ANNUAL'
    ]

    # Define fields that need overrides and scaling
    # Format: (output_column_name, field_name, overrides_dict, scale_factor)
    override_fields = [
        # 100-day average volume
        ('AVAT_100d', 'INTERVAL_AVG', {
            'CALC_INTERVAL': '100D',
            'MARKET_DATA_OVERRIDE': 'TURNOVER',
            'CRNCY': 'USD',
            'END_DATE_OVERRIDE': yesterday,
            'PERIODICITY_OVERRIDE': 'D'
        }, 1),

        # 20-day average volume
        ('AVAT_20d', 'INTERVAL_AVG', {
            'CALC_INTERVAL': '20D',
            'MARKET_DATA_OVERRIDE': 'TURNOVER',
            'CRNCY': 'USD',
            'END_DATE_OVERRIDE': yesterday,
            'PERIODICITY_OVERRIDE': 'D'
        }, 1),

        # Cash and equivalents (quarterly, scale by 1M)
        ('CASH_AND_EQUIVS', 'CASH_CASH_EQTY_STI_DETAILED', {
            'FUND_PER': 'Q'
        }, 1000000),

        # Short-term borrowings (quarterly, scale by 1M)
        ('BS_ST_BORROW', 'BS_ST_BORROW', {
            'FUND_PER': 'Q'
        }, 1000000),

        # Short-term lease liabilities (quarterly, scale by 1M)
        ('BS_ST_LEASE_LIAB', 'ST_CAPITALIZED_LEASE_LIABILITIES', {
            'FUND_PER': 'Q'
        }, 1000000),

        # Long-term borrowings (quarterly, scale by 1M)
        ('BS_LT_BORROW', 'BS_LT_BORROW', {
            'FUND_PER': 'Q'
        }, 1000000),

        # Long-term lease liabilities (quarterly, scale by 1M)
        ('LT_LEASES', 'LT_CAPITALIZED_LEASE_LIABILITIES', {
            'FUND_PER': 'Q'
        }, 1000000),

        # Cash flow from operations - quarterly (scale by 1M)
        ('CFO_Q', 'CF_CASH_FROM_OPER', {
            'FUND_PER': 'Q'
        }, 1000000),

        # Cash flow from operations - TTM (scale by 1M)
        ('CFO_TTM', 'TRAIL_12M_CASH_FROM_OPER', {
            'FUND_PER': 'Q'
        }, 1000000),

        # Free cash flow - TTM (scale by 1M)
        ('FCF_TTM', 'TRAIL_12M_FREE_CASH_FLOW', {
            'FUND_PER': 'Q'
        }, 1000000),

        # Free cash flow - quarterly (scale by 1M)
        ('FCF_Q', 'CF_FREE_CASH_FLOW', {
            'FUND_PER': 'Q'
        }, 1000000),

        # Equity float (scale by 1M) - Note: This may need BDH instead of BDP
        ('EQY_FLOAT2', 'EQY_FLOAT', {}, 1000000),

        # Interval high with 60-day lookback
        ('INTERVAL_HIGH_60D', 'INTERVAL_HIGH', {
            'START_DATE_OVERRIDE': sixty_days_ago
        }, 1),
    ]

    # Create collector instance
    collector = BloombergDataCollector()

    # Connect to Bloomberg
    if collector.connect():
        # Fetch standard fields
        print("\n=== Fetching standard fields ===")
        df_standard = collector.get_reference_data(tickers, standard_fields)

        # Dictionary to store all override field dataframes
        override_dfs = {}

        # Fetch each override field separately
        for col_name, field_name, overrides, scale in override_fields:
            print(f"\n=== Fetching {col_name} ===")
            try:
                df_temp = collector.get_reference_data(
                    tickers,
                    [field_name],
                    overrides=overrides if overrides else None
                )

                if df_temp is not None:
                    # Scale the values if needed
                    if scale != 1:
                        df_temp[field_name] = pd.to_numeric(df_temp[field_name], errors='coerce') * scale

                    # Rename column to the desired output name
                    df_temp.rename(columns={field_name: col_name}, inplace=True)
                    override_dfs[col_name] = df_temp
                else:
                    logger.warning(f"Failed to fetch {col_name}")
            except Exception as e:
                logger.error(f"Error fetching {col_name}: {str(e)}")

        # Combine all dataframes
        if df_standard is not None:
            df_final = df_standard.copy()

            # Add each override field
            for col_name, df_override in override_dfs.items():
                df_final = pd.concat([df_final, df_override], axis=1)

            # ===================================================================
            # PHASE 2: Fetch dependent fields that need values from Phase 1
            # ===================================================================
            print("\n" + "=" * 80)
            print("PHASE 2: Fetching dependent fields using retrieved dates")
            print("=" * 80)

            # These fields need the 10-K date (LATEST_ANN_DT_ANNUAL) from the first query
            dependent_dfs = {}

            for ticker in tickers:
                try:
                    # Get the 10-K date for this ticker (if available)
                    if 'LATEST_ANN_DT_ANNUAL' in df_final.columns:
                        ann_date = df_final.loc[ticker, 'LATEST_ANN_DT_ANNUAL']

                        if pd.notna(ann_date):
                            # Convert date to string format YYYYMMDD if needed
                            if isinstance(ann_date, pd.Timestamp):
                                ann_date_str = ann_date.strftime('%Y%m%d')
                            elif isinstance(ann_date, str):
                                # Try to parse and reformat if needed
                                try:
                                    ann_date_str = pd.to_datetime(ann_date).strftime('%Y%m%d')
                                except:
                                    ann_date_str = ann_date.replace('-', '').replace('/', '')
                            else:
                                ann_date_str = str(ann_date).replace('-', '').replace('/', '')

                            print(f"\n  {ticker}: Using 10-K date = {ann_date_str}")

                            sixty_days_before_ann = (ann_date - timedelta(days=60)).strftime('%Y%m%d')
                            sixty_days_before_ann_str = pd.to_datetime(sixty_days_before_ann).strftime('%Y%m%d')

                            # INTERVAL_HIGH1: from (10-K date - 60d to 10-K date)
                            print(f"    Fetching INTERVAL_HIGH1 (60d before 10-K to 10-K)...")
                            df_ih1 = collector.get_reference_data(
                                [ticker],
                                ['INTERVAL_HIGH'],
                                overrides={
                                    'START_DATE_OVERRIDE': sixty_days_before_ann_str,
                                    'END_DATE_OVERRIDE': ann_date_str
                                }
                            )
                            if df_ih1 is not None:
                                df_ih1.rename(columns={'INTERVAL_HIGH': 'INTERVAL_HIGH1'}, inplace=True)
                                if ticker not in dependent_dfs:
                                    dependent_dfs[ticker] = {}
                                dependent_dfs[ticker]['INTERVAL_HIGH1'] = df_ih1.loc[ticker, 'INTERVAL_HIGH1']

                            # INTERVAL_HIGH2: from 10-K date to now (no end date)
                            print(f"    Fetching INTERVAL_HIGH2 (from 10-K date)...")
                            df_ih2 = collector.get_reference_data(
                                [ticker],
                                ['INTERVAL_HIGH'],
                                overrides={
                                    'START_DATE_OVERRIDE': ann_date_str
                                }
                            )
                            if df_ih2 is not None:
                                df_ih2.rename(columns={'INTERVAL_HIGH': 'INTERVAL_HIGH2'}, inplace=True)
                                if ticker not in dependent_dfs:
                                    dependent_dfs[ticker] = {}
                                dependent_dfs[ticker]['INTERVAL_HIGH2'] = df_ih2.loc[ticker, 'INTERVAL_HIGH2']

                            # EQUITY_FLOAT1: Equity float as of the 10-K filing date (scaled by 1M)
                            # Use BDH (historical) instead of BDP to get point-in-time data
                            print(f"    Fetching EQUITY_FLOAT1 (as of 10-K date) using BDH...")
                            df_eqf1 = collector.get_historical_data(
                                [ticker],
                                ['EQY_FLOAT'],
                                start_date=ann_date_str,
                                end_date=ann_date_str
                            )
                            if df_eqf1 is not None and 'EQY_FLOAT' in df_eqf1.columns:
                                # Scale by 1 million
                                equity_float_value = pd.to_numeric(df_eqf1.loc[ticker, 'EQY_FLOAT'], errors='coerce')
                                if pd.notna(equity_float_value):
                                    equity_float_value = equity_float_value * 1000000
                                if ticker not in dependent_dfs:
                                    dependent_dfs[ticker] = {}
                                dependent_dfs[ticker]['EQY_FLOAT1'] = equity_float_value
                            else:
                                logger.warning(f"    No historical EQY_FLOAT data for {ticker} on {ann_date_str}")
                        else:
                            logger.warning(f"  {ticker}: No 10-K date available, skipping dependent fields")
                    else:
                        logger.warning("LATEST_ANN_DT_ANNUAL not found in dataframe")

                except Exception as e:
                    logger.error(f"Error fetching dependent fields for {ticker}: {str(e)}")

            # Add dependent fields to final dataframe
            if dependent_dfs:
                df_dependent = pd.DataFrame.from_dict(dependent_dfs, orient='index')
                df_final = pd.concat([df_final, df_dependent], axis=1)

            # Reset index to turn tickers into a column
            df_final = df_final.reset_index()
            df_final.rename(columns={'index': 'Ticker'}, inplace=True)

            # Define desired column order
            column_order = [
                'NAME',
                'INDUSTRY_GROUP',
                'PX_LAST',
                # 'EQY_INIT_PO_DT',
                '3MO_CALL_IMP_VOL',  # 3MO_CALL_IMP_VOL
                '12MO_CALL_IMP_VOL',  # 12MO_CALL_IMP_VOL
                'VOLATILITY_90D',
                'CUR_MKT_CAP',
                'AVAT_100d',
                'AVAT_20d',
                'CASH_AND_EQUIVS',
                'BS_ST_BORROW',
                'BS_ST_LEASE_LIAB',
                'ST_DEBT',
                'BS_LT_BORROW',
                'LT_LEASES',
                'LT_DEBT',
                'CFO_Q',
                'CFO_TTM',
                'FCF_Q',
                'FCF_TTM',
                'EQY_FLOAT1',
                'EQY_FLOAT2',
                'INTERVAL_HIGH_60D',
                'INTERVAL_HIGH1',
                'INTERVAL_HIGH2',
                'MOST_RECENT_PERIOD_END_DT',
                'LATEST_ANN_DT_QTRLY',
                'OFFERING_PRELIM_FILING_DT',
                'LATEST_ANN_DT_ANNUAL'
            ]

            # Bring in all worksheet columns
            print("Calculating additional fields...")
            df_final['ST_DEBT'] = df_final['BS_ST_BORROW'] - df_final['BS_ST_LEASE_LIAB']
            df_final['LT_DEBT'] = df_final['BS_LT_BORROW'] - df_final['LT_LEASES']
            df_final['FloatValue_preFile'] = df_final['EQY_FLOAT1'] * df_final['INTERVAL_HIGH1']
            df_final['FloatValue_postFile'] = df_final['EQY_FLOAT2'] * df_final['INTERVAL_HIGH2']
            df_final['FloatValue_last60d'] = df_final['EQY_FLOAT2'] * df_final['INTERVAL_HIGH_60D']

            df_final['baby_shelf_filter'] = (
                    (df_final['FloatValue_preFile'] < 75000000) & (df_final['FloatValue_postFile'] < 75000000) |
                    df_final[
                        'FloatValue_preFile'].isna() | df_final['FloatValue_postFile'].isna()).astype(int)
            df_final['shelf_limit'] = np.where(
                df_final['baby_shelf_filter'].isna() | df_final['FloatValue_last60d'].isna(),
                '?',
                np.where(df_final['baby_shelf_filter'] == 1,
                         (1 / 3) * np.maximum(df_final['FloatValue_preFile'], df_final['FloatValue_last60d']),
                         'Unlimited')
            )

            # Burn rate columns
            df_final['Burn_Q'] = np.where((df_final['CASH_AND_EQUIVS'] - df_final['ST_DEBT']) > 0,
                                          3 * (df_final['CASH_AND_EQUIVS'] - df_final['ST_DEBT']) / -df_final['FCF_Q'],
                                          'Net Debt')
            df_final['Burn_TTM'] = np.where((df_final['CASH_AND_EQUIVS'] - df_final['ST_DEBT']) > 0,
                                            12 * (df_final['CASH_AND_EQUIVS'] - df_final['ST_DEBT']) / -df_final[
                                                'FCF_TTM'], 'Net Debt')

            df_final['Burn_Q'] = pd.to_numeric(df_final['Burn_Q'], errors='coerce').fillna('')
            df_final['Burn_TTM'] = pd.to_numeric(df_final['Burn_TTM'], errors='coerce').fillna('')

            df_final['AVAT_100d_Burn_Q'] = -1 * df_final['FCF_Q'] / (63 * df_final['AVAT_100d'])
            df_final['AVAT_20d_Burn_Q'] = -1 * df_final['FCF_Q'] / (63 * df_final['AVAT_20d'])
            df_final['AVAT_100d_Burn_TTM'] = -1 * df_final['FCF_TTM'] / (252 * df_final['AVAT_100d'])
            df_final['AVAT_20d_Burn_TTM'] = -1 * df_final['FCF_TTM'] / (252 * df_final['AVAT_20d'])

            # Create empty columns to mimic original sheet
            # df_final['NAME'] = ""
            # df_final['INDUSTRY_GROUP'] = ""
            df_final['FLAG'] = ""
            df_final['Assignment'] = ""
            df_final['eloc_filter'] = ""
            df_final['atm_filter'] = ""
            df_final['Notes'] = ""

            # Date manipulations - keep as datetime for proper Excel formatting
            df_final['10-K_Date'] = pd.to_datetime(df_final['LATEST_ANN_DT_ANNUAL'], errors='coerce')
            df_final['10-K_Date_minus60d'] = pd.to_datetime(df_final['LATEST_ANN_DT_ANNUAL'],
                                                            errors='coerce') - pd.Timedelta(days=60)

            # Ensure all date columns are datetime objects
            date_cols_to_convert = ['MOST_RECENT_PERIOD_END_DT', 'LATEST_ANN_DT_QTRLY',
                                    'OFFERING_PRELIM_FILING_DT', 'LATEST_ANN_DT_ANNUAL']
            for col in date_cols_to_convert:
                if col in df_final.columns:
                    df_final[col] = pd.to_datetime(df_final[col], errors='coerce')

            # Apply filters
            print("Applying filters...")
            filters = apply_filters(df_final)

            # Add test columns to dataframe
            for test_name, test_result in filters.items():
                df_final[test_name] = test_result

            test_cols = list(filters.keys())

            # Create fail reason column
            df_final['fail_reason'] = create_fail_reason(df_final, test_cols)

            # Determine which rows pass all tests
            df_final['final_pass'] = df_final[test_cols].all(axis=1)

            print(f"✓ Filters applied: {df_final['final_pass'].sum()} passing rows out of {len(df_final)}")

            # Output 1: Full results with all test columns
            output_cols = ['Ticker'] + test_cols + ['fail_reason']
            df_output = df_final[output_cols].copy()

            output_path_full = f"full_filter_results.xlsx"
            df_output.to_excel(output_path_full, index=False, engine='openpyxl')
            format_excel_columns(output_path_full)
            print(f"✅ Full results: {output_path_full}")

            # Output 2: Only passing tickers
            df_good = df_final[df_final['final_pass']][['Ticker'] + test_cols].copy()

            output_path_good = f"good_tickers.xlsx"
            df_good.to_excel(output_path_good, index=False, engine='openpyxl')
            format_excel_columns(output_path_good)
            print(f"✅ Passing tickers: {output_path_good}")

            # Reorder columns (only include columns that exist)
            existing_columns = [col for col in column_order if col in df_final.columns]
            remaining_columns = [col for col in df_final.columns if col not in column_order]
            df_final = df_final[existing_columns + remaining_columns]

            # Print whole dataframe to mimic original datasheet
            col_order = ['NAME', 'INDUSTRY_GROUP', 'Ticker', 'test_adv_gt_50k', 'eloc_filter', 'atm_filter',
                         'baby_shelf_filter', 'shelf_limit', 'PX_LAST', '3MO_CALL_IMP_VOL', '12MO_CALL_IMP_VOL',
                         'VOLATILITY_90D', 'CUR_MKT_CAP', 'AVAT_100d', 'AVAT_20d', 'CASH_AND_EQUIVS', 'BS_ST_BORROW',
                         'BS_ST_LEASE_LIAB', 'ST_DEBT', 'BS_LT_BORROW', 'LT_LEASES', 'LT_DEBT', 'CFO_Q', 'CFO_TTM',
                         'FCF_Q',
                         'FCF_TTM', 'Assignment', 'FLAG', 'Notes', 'MOST_RECENT_PERIOD_END_DT', 'LATEST_ANN_DT_QTRLY',
                         'OFFERING_PRELIM_FILING_DT', 'Burn_Q', 'Burn_TTM', 'AVAT_100d_Burn_Q', 'AVAT_20d_Burn_Q',
                         'AVAT_100d_Burn_TTM', 'AVAT_20d_Burn_TTM', 'LATEST_ANN_DT_ANNUAL', '10-K_Date',
                         '10-K_Date_minus60d', 'EQY_FLOAT1', 'EQY_FLOAT2', 'INTERVAL_HIGH1', 'INTERVAL_HIGH2',
                         'INTERVAL_HIGH_60D', 'FloatValue_preFile', 'FloatValue_postFile', 'FloatValue_last60d']
            bbg_headers = ['NAME', 'INDUSTRY_GROUP', 'TICKER', 'Low Volume', 'ELOC', 'ATM', 'Baby Shelf?',
                           'Shelf Limit', 'PX_LAST', '3MO_CALL_IMP_VOL', '12MO_CALL_IMP_VOL', 'VOLATILITY_90d',
                           'CUR_MKT_CAP', 'AVAT (100d)', 'AVAT (20d)', 'CASH & EQUIVS', 'ST BORROWINGS', 'ST LEASES',
                           'ST DEBT', 'LT BORROW', 'LT LEASES', 'LT DEBT', 'CFO (Q)', 'CFO (TTM)', 'FCF (Q)',
                           'FCF (TTM)', 'Assignment', 'FLAG', 'Notes', 'MOST_RECENT_PERIOD_END_DT',
                           'LATEST_ANN_DT_QTRLY', 'OFFERING_PRELIM_FILING_DT', 'Burn_Q', 'Burn_TTM', 'AVAT (100d)',
                           'AVAT (20d)', 'AVAT (100d)', 'AVAT (20d)', 'LATEST_ANN_DT_ANNUAL', '10-K Date',
                           '10-K Date - 60d', 'EQY_FLOAT', 'EQY_FLOAT', 'INTERVAL_HIGH', 'INTERVAL_HIGH',
                           'INTERVAL_HIGH', 'Float Value', 'Float Value', 'Float Value']

            df_inOrder = df_final[col_order]
            df_inOrder.insert(0, '', '')
            # bbg_row=pd.DataFrame([bbg_headers],columns=df_inOrder)
            # df_inOrder = pd.concat([bbg_row, df_inOrder], ignore_index=True)

            # Output 3: Writing data sheet to re-create original data sheet
            whole_dataframe_file = f"spreadsheet_data.xlsx"
            df_inOrder.to_excel(whole_dataframe_file, index=False, engine='openpyxl')
            format_excel_columns(whole_dataframe_file)
            print(f"✅ Writing data sheet: {whole_dataframe_file}")

            # Display results
            print("\n" + "=" * 80)
            print("FINAL DATA SUMMARY")
            print("=" * 80)
            print(df_final)

            # Save to Excel
            output_file = f'bloomberg_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            df_final.to_excel(output_file, index=False, engine='openpyxl')
            format_excel_columns(output_file)
            logger.info(f"Data saved to {output_file}")

            # Display data info
            print(f"\nShape: {df_final.shape}")
            print(f"\nColumns: {list(df_final.columns)}")
            print(f"\nNull values:\n{df_final.isnull().sum()}")

            # Show statistics for numeric columns
            print("\nNumeric columns summary:")
            print(df_final.select_dtypes(include=['number']).describe())

        # Disconnect
        collector.disconnect()
    else:
        logger.error("Failed to connect to Bloomberg API")


if __name__ == "__main__":
    main()